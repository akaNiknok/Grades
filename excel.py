from __future__ import print_function
import openpyxl

TEST_LABEL_ROW = 7
TEST_TOTAL_ROW = 8

wb = openpyxl.load_workbook("help/SAMPLE-SHEET.xlsx", data_only=True)

ws = wb.get_sheet_by_name("Raw.Score-1st")

for row in range(1, ws.max_row):
    if ws.cell(row=row, column=1).value == 10:
        user_row = row

for column in range(3, ws.max_column):
    if ws.cell(row=TEST_LABEL_ROW, column=column).value is not None:
        print(ws.cell(row=TEST_LABEL_ROW, column=column).value, end=": ")
        print(ws.cell(row=user_row, column=column).value, end="/")
        print(ws.cell(row=TEST_TOTAL_ROW, column=column).value, end="\n")
