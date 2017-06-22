import os
import openpyxl
from openpyxl.utils import range_boundaries
from bs4 import BeautifulSoup


def read_htmls(grade, section, cn):
    """Read pre-rendered tables of every subject of a student and return a dict
    Return format:
        subjects = {
            "Subject":
                "Trimester": pre_rendered table
        }"""

    # Get file directory using users grade and section
    filedir = "excels/{}/{}/".format(grade, section)

    # Get all files (subjects) in file directory
    try:
        files = os.listdir(filedir)
    except OSError:
        return "Empty"

    subjects = {}

    # Loop through subjects
    for file in files:

        # only read .html.j2 files
        if file.endswith(".j2"):

            trimesters = {}
            incl_rows = [0, 1, 2, 3, cn + 3]  # Headers and user row

            # Open the pre-rendered html
            with open("excels/{}/{}/{}".format(grade, section, file)) as f:
                soup = BeautifulSoup(f.read().decode("utf-8"))

            # Loop through trimesters
            for div in soup.find_all("div"):

                # Get rows and clear table
                rows = div.find_all("tr")
                div.table.clear()

                # Add headers and user row to new table
                for row in incl_rows:
                    div.table.append(rows[row])

                # Store the div in the trimesters
                trimesters[div.get("id")] = str(div)

            # Store the trimester in the subjects
            # And also get the filename as the key
            subjects[file.split(".")[0]] = trimesters

    return subjects


def read_excel(grade, section, subject):
    """Reads an excel file and returns a table list"""

    # Open excel file
    wb = openpyxl.load_workbook("excels/{}/{}/{}".format(grade,
                                                         section,
                                                         subject + ".xlsx"),
                                data_only=True)

    trimesters = {}

    # Loop through sheets
    for ws in wb.worksheets:

        if "Raw.Score" in ws.title:

            # Get boundary (min_col, min_row, max_col, max_row) of merged_cells
            merged_cells = [range_boundaries(r) for r in ws.merged_cell_ranges]

            table = []
            empty_cols = range(0, ws.max_column - 1)

            # Loop through rows
            for row in range(5, ws.max_row):

                # Contains columns (value, colspan) here
                new_row = []

                # Create an iterator object to be able to skip merged cells
                columns = iter(range(1, ws.max_column))

                # Loop through columns
                for column in columns:

                    # Get value of cell
                    value = ws.cell(row=row, column=column).value

                    # Check if the cell is not empty
                    if value is not None:
                        colspan = 1

                        # Remove column from empty_cols
                        if (column - 1) in empty_cols:
                            empty_cols.remove(column - 1)

                        # Check if cell is merged
                        for r in merged_cells:
                            if (r[0] == column) and (r[1] == row):
                                colspan += (r[2] - r[0])  # Add to colspan
                                break

                        # Append column's row and colspan to row
                        new_row.append((value, colspan))

                        # If cell is merged, skip (colspan - 1) iterations
                        if colspan != 1:
                            for x in range(colspan - 1):
                                next(columns)
                    else:
                        new_row.append(("", 1))

                # Append row to table
                table.append(new_row)

            # Remove empty columns
            for row in table:

                fake_col = 0  # To identify the index of the column in table
                real_col = 0  # To identify the column number
                remove_cols = []

                # Identify columns to be removed
                for col in row:
                    real_col += col[1]
                    fake_col += 1
                    if real_col in empty_cols:
                        remove_cols.append(fake_col)

                remove_cols.reverse()  # Reverse to escape IndexError

                # Remove the columns
                for col in remove_cols:
                    row.pop(col)

            # Store the table in the trimesters dictionary
            trimesters[ws.title.split("-")[1]] = table

    return trimesters
