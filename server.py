import os
import json
from flask import Flask, render_template, request, session
from flask import redirect, make_response, send_file
from flask_sqlalchemy import SQLAlchemy
import openpyxl
from openpyxl.utils import range_boundaries

app = Flask(__name__)
app.config.from_object("config")
db = SQLAlchemy(app)


class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(20), unique=True)
    password = db.Column(db.String(20))
    acc_type = db.Column(db.String(11))

    firstname = db.Column(db.String())
    middlename = db.Column(db.String())
    lastname = db.Column(db.String())

    grade = db.Column(db.Integer)
    section = db.Column(db.String(4))
    CN = db.Column(db.Integer)

    subject = db.Column(db.String(7))
    sections = db.Column(db.String)

    def __init__(self,
                 username,
                 password,
                 acc_type,
                 firstname,
                 middlename,
                 lastname):
        self.username = username
        self.password = password
        self.acc_type = acc_type
        self.firstname = firstname
        self.middlename = middlename
        self.lastname = lastname

    def mi(self):
        """Returns middle initial of user"""
        mi = ""

        for name in self.middlename.split():
            mi += name[0]
            mi += "."

        return mi


db.create_all()


@app.route('/', methods=["POST", "GET"])
def index():
    user_id = session.get("user_id")

    if request.method == "POST":

        # Refresh for students
        user = User.query.get(user_id)
        subjects = read_excels(user.grade, user.section, user.CN)
        session["subjects"] = subjects

        return redirect("/")

    else:
        # Check if logged in
        if user_id:
            user = User.query.get(user_id)

            # Get subjects from session if user is student
            if user.acc_type == "student":
                return render_template(
                    "index.html.j2",
                    user=user,
                    subjects=session.get("subjects")
                )

            # Get sections if user is teacher
            elif user.acc_type == "teacher":
                return render_template(
                    "index.html.j2",
                    user=user,
                    sections=json.loads(user.sections)
                )

            # Get teachers if user is coordinator
            elif user.acc_type == "coordinator":
                return render_template(
                    "index.html.j2",
                    user=user,
                    teachers=User.query.filter_by(
                        subject=user.subject,
                        acc_type="teacher"
                    ).all(),
                    loads=json.loads
                )
        else:
            user = None

        return render_template("index.html.j2", user=user)


@app.route('/about')
def about():
    user_id = session.get("user_id")

    if user_id:
        user = User.query.get(user_id)
    else:
        user = None

    return render_template("about.html.j2", user=user)


@app.route('/register', methods=["POST", "GET"])
def register():
    if request.method == "POST":

        # Get data from general forms
        username = request.form["username"]
        password = request.form["password"]
        re_password = request.form["re_password"]
        acc_type = request.form["acc_type"]

        firstname = request.form["firstname"]
        middlename = request.form["middlename"]
        lastname = request.form["lastname"]

        # Student Specific Forms
        grade = request.form["grade"]
        section = request.form["section"]
        CN = request.form["CN"]

        # Teacher Specific Forms
        new_teacher_pass = request.form["new_teacher_pass"]
        new_teacher_subject = request.form["new_teacher_subject"]

        # Coordinator Specific Forms
        new_coord_pass = request.form["new_coord_pass"]
        new_coord_subject = request.form["new_coord_subject"]

        # Check if username is taken
        if User.query.filter_by(username=username).first() is None:

            # Validate retype password
            if password == re_password:

                # Add grade, section and CN when account type is student
                if acc_type == "student":
                    student = User(username,
                                   password,
                                   acc_type,
                                   firstname,
                                   middlename,
                                   lastname)
                    student.grade = int(grade)
                    student.section = section
                    student.CN = int(CN)

                    db.session.add(student)

                # Add subject and section when account type is teacher
                elif (acc_type == "teacher" and
                        new_teacher_pass == "CSQC new teach"):
                    teacher = User(username,
                                   password,
                                   acc_type,
                                   firstname,
                                   middlename,
                                   lastname)
                    teacher.subject = new_teacher_subject
                    teacher.sections = "[]"
                    db.session.add(teacher)

                # Add subject when account type is coordinator
                elif (acc_type == "coordinator" and
                        new_coord_pass == "CSQC new coord"):
                    coord = User(username,
                                 password,
                                 acc_type,
                                 firstname,
                                 middlename,
                                 lastname)
                    coord.subject = new_coord_subject
                    db.session.add(coord)

                else:
                    # Entered the wrong new_*_pass
                    return render_template("register.html.j2", error=acc_type)

                db.session.commit()

            else:
                # Entered wrong retype password
                return render_template("register.html.j2", error="password")
        else:
            # Username taken
            return render_template("register.html.j2", error="username")

        # Login new user if "All Izz Well" :D
        user = User.query.filter_by(username=username).first()
        response = make_response(redirect("/"))
        session["user_id"] = user.id

        return response
    else:
        return render_template("register.html.j2")


@app.route('/login', methods=["POST", "GET"])
def login():
    if request.method == "POST":

        # Get data from form
        username = request.form["username"]
        password = request.form["password"]

        user = User.query.filter_by(username=username).first()

        # Validate username
        if user is None:
            return render_template("login.html.j2", error="username")

        # Validate password
        elif user.password == password:
            response = make_response(redirect("/"))
            session["user_id"] = user.id
            return response
        else:
            return render_template("login.html.j2", error="password")

    else:
        return render_template("login.html.j2")


@app.route('/logout')
def logout():
    session.clear()
    return redirect("/")


@app.route('/user/<username>', methods=["POST", "GET"])
def user(username):
    user_id = session.get("user_id")

    if request.method == "POST":

        # Get user
        user = User.query.get(user_id)

        # Check if user wants to change password or delete his/her account
        submit = request.form["submit"]

        if submit == "Change Password":

            # Get data from form
            orig_pass = request.form["orig_pass"]
            new_pass = request.form["new_pass"]
            re_new_pass = request.form["re_new_pass"]

            # Check if original password matches user's password
            if orig_pass == user.password:

                # Check if new password matches retype password
                if new_pass == re_new_pass:

                    # Change the password and save to database
                    user.password = new_pass
                    db.session.commit()

                    return render_template("user.html.j2",
                                           user=user,
                                           view_user=user,
                                           success=True)
                else:
                    return render_template("user.html.j2",
                                           user=user,
                                           view_user=user,
                                           error="re_new_pass")
            else:
                return render_template("user.html.j2",
                                       user=user,
                                       view_user=user,
                                       error="orig_pass")

        else:
            # Get data from form
            username = request.form["username"]
            password = request.form["password"]

            # Validate username
            if user.username != username:
                return render_template("user.html.j2",
                                       user=user,
                                       view_user=user,
                                       error="username")
            # Validate password
            elif user.password != password:
                return render_template("user.html.j2",
                                       user=user,
                                       view_user=user,
                                       error="password")

            else:
                # Logout and Delete account
                session.clear()
                db.session.delete(user)
                db.session.commit()

                return redirect("/")

    else:
        if user_id:
            user = User.query.get(user_id)

            # Check if logged in user is viewing his own account
            if user.username == username:
                view_user = user
            else:
                view_user = User.query.filter_by(username=username).first()
        else:
            user = None
            view_user = User.query.filter_by(username=username).first()

        return render_template("user.html.j2", user=user, view_user=view_user)


@app.route('/excels/<grade>/<section>/<subject>', methods=["POST", "GET"])
def excels(grade, section, subject):
    if request.method == "POST":

        # Refresh for teachers and coordinators
        session["table"] = read_excel(grade, section, subject)
        return redirect("/excels/{}/{}/{}".format(grade, section, subject))

    else:
        # Get user
        user_id = session.get("user_id")

        if user_id:
            user = User.query.get(user_id)
        else:
            return redirect("/")

        # Only allow teachers and coordinators to view excels
        if user.acc_type != "teacher" and user.acc_type != "coordinator":
            return redirect("/")

        return render_template("excel.html.j2",
                               user=user,
                               grade=grade,
                               section=section,
                               subject=subject,
                               table=session.get("table"))


@app.route('/upload', methods=["POST", "GET"])
def upload():
    user_id = session.get("user_id")

    if request.method == "POST":
        file = request.files['file']

        # Check if the file is present and is a spreadsheet (.xlsx, .xls)
        if file and (file.filename.endswith(".xlsx")
                     or file.filename.endswith(".xls")):

            # Get the user
            user = User.query.get(user_id)

            # Rename the file to subject.xlsx
            filename = user.subject + ".xlsx"

            # Directory for the excel file
            filedir = "excels/{}/{}/".format(request.form["grade"],
                                             request.form["section"])

            # If the file already exist, don't add anonther section to the list
            if not os.path.isfile(filedir + filename):
                sections = json.loads(user.sections)

                # Add the section as tuple to the list
                sections.append((request.form["grade"],
                                 request.form["section"]))

                # Save the list
                user.sections = json.dumps(sections)
                db.session.commit()

            # Create Directory if it does not exist
            if not os.path.isdir(filedir):
                os.makedirs(filedir)

            # Save the file in the correct directory
            file.save(os.path.join(filedir, filename))

            return render_template("upload.html.j2", success=True)
        else:
            return render_template("upload.html.j2", error=True)

    else:

        # Require user to be a teacher or coordinator
        if user_id:
            user = User.query.get(user_id)
            if user.acc_type == "teacher" or "coordinator":
                return render_template("upload.html.j2", user=user)
            else:
                return redirect("/")
        else:
            return redirect("/")


@app.route('/download/<grade>/<section>/<subject>')
def download(grade, section, subject):
    return send_file("excels/{}/{}/{}.xlsx".format(grade, section, subject),
                     as_attachment=True)


@app.route('/delete', methods=["POST"])
def delete():

    # Get user
    user_id = session.get("user_id")
    user = User.query.get(user_id)

    # Get index
    delete_index = int(request.form["delete_index"])

    # Load the sections list
    sections = json.loads(user.sections)
    section = sections[delete_index]

    # Remove the section from the list and delete the file
    sections.pop(delete_index)
    os.remove(
        "excels/{}/{}/{}.xlsx".format(section[0], section[1], user.subject)
    )

    # Save the list
    user.sections = json.dumps(sections)
    db.session.commit()

    return redirect("/")


def read_excels(grade, section, cn):
    """Return format:
        subjects = {
            "Subject": {
                "Test": (int(Student Score), int(Total Score))
            }
        }
    """

    # Rows for label and total score
    TEST_LABEL_ROW = 7
    TEST_TOTAL_ROW = 8

    # Get file directory using users grade and section
    filedir = "excels/{}/{}/".format(grade, section)

    # Get all files (subjects) in file directory
    try:
        files = os.listdir(filedir)
    except OSError:
        return None

    subjects = {}

    # Loop per subject
    for file in files:

        # Open the excel file
        wb = openpyxl.load_workbook(os.path.join(filedir, file),
                                    data_only=True)

        # Open the sheet (subject to change)
        ws = wb.worksheets[0]

        # Find CN rows
        for row in range(1, ws.max_row):
            if ws.cell(row=row, column=1).value == cn:
                user_row = row

        Tests = {}

        # Store tests in dictionary
        for col in range(3, ws.max_column):

            test_label = ws.cell(row=TEST_LABEL_ROW, column=col).value
            student_score = ws.cell(row=user_row, column=col).value
            total_score = ws.cell(row=TEST_TOTAL_ROW, column=col).value

            # Only include scores with label
            if ((test_label not in (None, "TS", "PS", "EP"))
                    and (student_score is not None)):

                if test_label in Tests:
                    Tests[test_label][0] += student_score
                    Tests[test_label][1] += total_score
                else:
                    Tests[test_label] = [student_score, total_score]

        # Store the tests in subject
        # Also removes the file extension
        subjects[os.path.splitext(file)[0]] = Tests

    return subjects


def read_excel(grade, section, subject):
    """Reads an excel file and returns a table list"""

    # Open excel file
    wb = openpyxl.load_workbook("excels/{}/{}/{}".format(grade,
                                                         section,
                                                         subject + ".xlsx"),
                                data_only=True)

    # Open the sheet (subject to change)
    ws = wb.worksheets[0]

    # Get boundaries (min_col, min_row, max_col, max_row) of merged_cells
    merged_cells = [range_boundaries(r) for r in ws.merged_cell_ranges]

    table = []

    # Loop through rows
    for row in range(5, ws.max_row):

        # Contains columns (value, colspan) here
        new_row = []

        # Create an iterator object to be able to skip merged cells
        columns = iter(range(1, ws.max_column))

        # Loop through columns
        for column in columns:

            # Get value of cell
            value = ws.cell(row=row, column=column).value

            # Check if the cell is not empty
            if value is not None:
                colspan = 1

                # Check if cell is merged
                for r in merged_cells:
                    if (r[0] == column) and (r[1] == row):
                        colspan += (r[2] - r[0])  # Add to colspan
                        break

                new_row.append((value, colspan))

                # If cell is merged, skip (colspan - 1) iterations
                if colspan != 1:
                    for x in range(colspan - 1):
                        next(columns)
            else:
                new_row.append(("", 1))

        table.append(new_row)

    return table


if __name__ == "__main__":
    app.run(host="0.0.0.0", threaded=True)


# At World's End (POTC :stuck_out_tongue_closed_eyes::skull:)
